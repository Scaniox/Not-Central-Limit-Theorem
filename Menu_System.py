from matplotlib.transforms import Transform
import openpyxl as xl
from collections import Counter
from scipy.stats import norm
import numpy as np
import matplotlib.pyplot as plt
import pygame as pg
from pathlib import Path
import io
import random as rng
import Menu_Sprites as MS

vec2 = pg.math.Vector2
default_rect = lambda : pg.rect.Rect(0,0,10,10)

# convert plot to surface
def plot_to_surface(fig):
    fig.canvas.draw()

    io_buf = io.BytesIO()
    fig.savefig(io_buf, format='raw', dpi=100)
    io_buf.seek(0)
    img_arr = np.reshape(np.frombuffer(io_buf.getvalue(), dtype=np.uint8),
                        newshape=(int(fig.bbox.bounds[3]), int(fig.bbox.bounds[2]), -1))
    io_buf.close()

    img = pg.surfarray.make_surface(img_arr[:,::-1,0:3])
    return pg.transform.rotate(img,270)


# take samples and count averages
def sample_averages(data, sample_size, sample_count):
    avgs = []
    for sample in range(sample_count):
        avgs.append(sum(rng.choices(data, k=sample_size))/sample_size)
    return avgs
    

class Ui_Screen():
    def __init__(self, game):
        self.game = game

        self.elements = pg.sprite.Group()
        self.background = False

        self.screen = self.game.screen

    def tick(self, event_list, dt):
        self.elements.update(dt, event_list)
        if self.background:
            self.game.screen.blit(self.bg_img, self.bg_rect)
        self.elements.draw(self.game.screen)

    def rescale(self):
        screen_rect = pg.rect.Rect(0, 0, *self.screen.get_size())
        screen_size = vec2(screen_rect.size)

        # rescale the background if it is present
        if self.background:
            # choose scale factor to fill screen
            match_width_SF = screen_rect.width / self.background.get_width()
            match_height_SF = screen_rect.height / self.background.get_height()
            scale_factor = max(match_height_SF, match_width_SF)
            
            # rescale image
            background_size = vec2(self.background.get_size()) * scale_factor
            self.bg_img = pg.transform.scale(self.background, background_size)
            
            # position image
            self.bg_rect = self.bg_img.get_rect()
            self.bg_rect.center = screen_rect.center

        # rescale all elements
        for elements in self.elements:
            elements.rescale()

class Main(Ui_Screen):
    def __init__(self, game):
        super().__init__(game)

        # init elements
        self.sample_size_slider = MS.Slider(game, default_rect())
        self.sample_size_slider.val = 0.5
        self.elements.add(self.sample_size_slider)
        self.sample_size_txt = MS.Text(game, default_rect(), "sample size: 0")
        self.elements.add(self.sample_size_txt)

        self.sample_count_slider = MS.Slider(game, default_rect())
        self.sample_count_slider.val = 0.5
        self.elements.add(self.sample_count_slider)
        self.sample_count_txt = MS.Text(game, default_rect(), "sample count: 0")
        self.elements.add(self.sample_count_txt)
        
        self.box_toggle = MS.Toggle(game, default_rect())
        self.box_toggle_txt = MS.Text(game, default_rect(), "bar chart")
        self.box_toggle.ticked = True
        self.elements.add(self.box_toggle)
        self.elements.add(self.box_toggle_txt)
        
        self.poly_toggle = MS.Toggle(game, default_rect())
        self.poly_toggle_txt = MS.Text(game, default_rect(), "frequency polygon")
        self.poly_toggle.ticked = True
        self.elements.add(self.poly_toggle)
        self.elements.add(self.poly_toggle_txt)

        self.expected_norm = MS.Toggle(game, default_rect())
        self.expected_norm_txt = MS.Text(game, default_rect(), "expected distribution")
        self.elements.add(self.expected_norm)
        self.elements.add(self.expected_norm_txt)

        self.gh_txt = MS.Text(game, default_rect(), "https://github.com/Scaniox")
        self.elements.add(self.gh_txt)

        # import data from xlsx file
        try:
            sheet = xl.load_workbook(game.config.xlsx_path).active
        except Exception as error:
            print(f"Error: {error}\n\nThis means it couldn't find the spreadsheet file\nmake sure it is in the same folder as this program and called data.xlsx")
            input()
        
        self.data = {}
        for row in sheet.iter_rows(1, sheet.max_row):
            self.data[row[0].value] = row[1].value

        self.pop_mean = np.mean(list(self.data.values()))
        self.pop_var = np.var(list(self.data.values()))

        self.x_min = min(self.data.values())
        self.x_max = max(self.data.values())
        self.y_max = norm.pdf(self.pop_mean,
                             loc=self.pop_mean,
                             scale=(self.pop_var/100)**0.5)


        # start plot
        self.fig, self.ax = plt.subplots()

        # setup 
        self.sample_size = 6
        self.sample_count = 5000 
        self.prev_states = [0,0, False, False, False, False, False]

        self.refresh_plot()

        # call rescale to render image
        self.rescale()

    def tick(self, event_list, dt):
        new_states = [self.sample_size_slider.val*99+1,
                      self.sample_count_slider.val*14999+1,
                      self.sample_size_slider.grabbed,
                      self.sample_count_slider.grabbed,
                      self.box_toggle.ticked,
                      self.poly_toggle.ticked,
                      self.expected_norm.ticked]

        if new_states[0:2] != self.prev_states[0:2]:
            self.sample_size_txt.text = f"sample size: {round(new_states[0])}"
            self.sample_count_txt.text = f"sample count: {round(new_states[1])}"
            self.sample_size_txt.rescale()
            self.sample_count_txt.rescale()

            self.sample_size = round(new_states[0])
            self.sample_count = round(new_states[1])

        if new_states[4:7] != self.prev_states[4:7] or \
           not new_states[2] and self.prev_states[2] or\
           not new_states[3] and self.prev_states[3]:

           self.refresh_plot()

        self.prev_states = new_states


        fig_height = self.game.screen.get_size()[1] * 3/4
        fig_width = self.fig_img.get_size()[0] * fig_height/ self.fig_img.get_size()[1]
        #fig_img = pg.transform.scale(fig_img, (fig_width, fig_height))
        fig_rect = self.fig_img.get_rect()
        fig_rect.center = (self.game.screen.get_size()[0]/2, self.game.screen.get_size()[1]*3/8)

        self.game.screen.fill((255,255,255))
        self.game.screen.blit(self.fig_img, fig_rect)

        # call parent tick function
        super().tick(event_list, dt)
    
    def refresh_plot(self):
        self.ax.clear()
        self.ax.set_ylim([0,self.y_max])
        self.ax.set_xlim([self.x_min,self.x_max])
        self.ax.set_ylabel("$P(\\overline{X}=x)$ $density$")
        self.ax.set_xlabel("x")

        sample_means = sample_averages(list(self.data.values()), self.sample_size, self.sample_count)
        freqs = Counter([round(sm, 10) for sm in sample_means])
        probs = [[bar,count*((self.sample_size))/self.sample_count/2] for bar, count in freqs.items()]
        keys, vals = zip(*sorted(probs))

        if self.box_toggle.ticked:
            self.ax.bar(keys, vals, width=0.05)
        if self.poly_toggle.ticked:
            self.ax.plot(keys, vals, color = "C1")
        if self.expected_norm.ticked:
            X = np.linspace(self.x_min,self.x_max, 200)
            Y = norm.pdf(X, loc=self.pop_mean, scale=(self.pop_var/self.sample_size)**0.5)
            self.ax.plot(X, Y, color="C2")

        # add text for mu, sigma and sample mean sigma
        sample_mean_var = np.var(list(sample_means))
        x_step = (self.x_max- self.x_min) / 3
        self.ax.text(0, 1.01,
                    f"population $\\mu$ : {round(self.pop_mean,3)}",
                    transform=self.ax.transAxes)
        self.ax.text(0.3333, 1.01,
                     f"population $\\sigma$ : {round(self.pop_var**0.5, 3)}",
                     transform=self.ax.transAxes)
        self.ax.text(0.6666, 1.01,
                     f"sample mean $\\sigma$ : {round(sample_mean_var**0.5, 3)}",
                     transform=self.ax.transAxes)

        self.fig_img = plot_to_surface(self.fig)

    def rescale(self):
        screen_rect = pg.rect.Rect(0, 0, *self.screen.get_size())
        screen_size = vec2(screen_rect.size)

        #  alignment
        #        c1   a1  c2      a2       c3  a3   c4 a4 c5
        #  r1    |========|================|========|=====|
        #  b1    |text: no|-------o------- | toggle:|  #  | b3
        #        |        |                |========|=====| r4
        #  r2    |========|================| toggle:|  #  | b4
        #  b2    |text: no|-------o------- |========|=====| r5
        #        |        |                | toggle:|  #  | b5
        #  r3    |========|================|========|=====|

        r1 = int(screen_size.y * 18/24)
        r2 = int(screen_size.y * 21/24)
        r3 = int(screen_size.y * 24/24)

        b1 = (r1 + r2)/2
        b2 = (r2 + r3)/2
        
        r4 = int(screen_size.y * 20/24)
        r5 = int(screen_size.y * 22/24)

        b3 = (r1 + r4)/2
        b4 = (r4 + r5)/2
        b5 = (r5 + r3)/2

        c1 = int(screen_size.x * 0/12)
        c2 = int(screen_size.x * 3/12)
        c3 = int(screen_size.x * 8/12)
        c4 = int(screen_size.x * 11/12)
        c5 = int(screen_size.x * 12/12)

        a1 = (c1 + c2)/2
        a2 = (c2 + c3)/2
        a3 = (c3 + c4)/2
        a4 = (c4 + c5)/2

        self.sample_size_txt.rect.height = (r2 - r1)/3
        self.sample_size_txt.rect.width = (c2 - c1)
        self.sample_size_txt.rect.centerx = a1
        self.sample_size_txt.rect.centery = b1

        self.sample_count_txt.rect.height = (r3 - r2)/3
        self.sample_count_txt.rect.width = (c2 - c1)
        self.sample_count_txt.rect.centerx = a1
        self.sample_count_txt.rect.centery = b2

        self.sample_size_slider.rect.height = (r2 - r1)/2
        self.sample_size_slider.rect.width = c3 - c2
        self.sample_size_slider.rect.centerx = a2
        self.sample_size_slider.rect.centery = b1
        
        self.sample_count_slider.rect.height = (r3 - r2)/2
        self.sample_count_slider.rect.width = c3 - c2
        self.sample_count_slider.rect.centerx = a2
        self.sample_count_slider.rect.centery = b2

        self.box_toggle_txt.rect.height = (r4 - r1)/2
        self.box_toggle_txt.rect.width = c4 - c3
        self.box_toggle_txt.rect.centerx = a3
        self.box_toggle_txt.rect.centery = b3
        
        self.poly_toggle_txt.rect.height = (r5 - r4)/2
        self.poly_toggle_txt.rect.width = c4 - c3
        self.poly_toggle_txt.rect.centerx = a3
        self.poly_toggle_txt.rect.centery = b4

        self.expected_norm_txt.rect.height = (r3 - r5)/2
        self.expected_norm_txt.rect.width = c4 - c3
        self.expected_norm_txt.rect.centerx = a3
        self.expected_norm_txt.rect.centery = b5
        
        self.box_toggle.rect.height = (r4 - r1)/2
        self.box_toggle.rect.width = (r4 - r1)/2
        self.box_toggle.rect.centerx = a4
        self.box_toggle.rect.centery = b3
        
        self.poly_toggle.rect.height = (r5 - r4)/2
        self.poly_toggle.rect.width = (r5 - r4)/2
        self.poly_toggle.rect.centerx = a4
        self.poly_toggle.rect.centery = b4

        self.expected_norm.rect.height = (r3 - r5)/2
        self.expected_norm.rect.width = (r3 - r5)/2
        self.expected_norm.rect.centerx = a4
        self.expected_norm.rect.centery = b5

        self.gh_txt.rect.height = screen_size[1] // 36
        self.gh_txt.rect.width = screen_size[0] // 6
        self.gh_txt.rect.left = 0 
        self.gh_txt.rect.bottom = screen_size[1]

        # call parent rescale method
        super().rescale()

