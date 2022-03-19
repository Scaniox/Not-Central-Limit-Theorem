import openpyxl as xl
from collections import Counter
from scipy.stats import norm
import numpy as np
import matplotlib.pyplot as plt
import pygame as pg
from pathlib import Path
import io
import random as rng
import Menu_Sprites as MS


# convert plot to surface
def plot_to_surface(fig):
    fig.canvas.draw()

    io_buf = io.BytesIO()
    fig.savefig(io_buf, format='raw', dpi=100)
    io_buf.seek(0)
    img_arr = np.reshape(np.frombuffer(io_buf.getvalue(), dtype=np.uint8),
                        newshape=(int(fig.bbox.bounds[3]), int(fig.bbox.bounds[2]), -1))
    io_buf.close()

    img = pg.surfarray.make_surface(img_arr[:,::-1,0:3])
    return pg.transform.rotate(img,270)


# take samples and count averages
def sample_averages(data, sample_size, sample_count):
    avgs = []
    for sample in range(sample_count):
        avgs.append(sum(rng.choices(data, k=sample_size))/sample_size)
    return avgs


# set up pygame window
pg.init()
screen = pg.display.set_mode((1000,800))

# import data from xlsx file
sheet = xl.load_workbook("Book3.xlsx").active
data = {}
for row in sheet.iter_rows(1, sheet.max_row):
    data[row[0].value] = row[1].value

pop_mean = np.mean(list(data.values()))
pop_var = np.var(list(data.values()))

# plot data
fig, ax = plt.subplots()

# setup 
sample_size = 1
sample_count = 10000 
box = True
f_poly = True
draw_norm = True

# ui setup
sample_size_slider = MS.Slider()

clock = pg.time.Clock()
loops = 0
running = True
while running:
    ax.clear()
    ax.set_ylim([0,1])
    ax.set_xlim([35,45])
    sample_means = sample_averages(list(data.values()), sample_size, sample_count)
    freqs = Counter([round(sm, 10) for sm in sample_means])
    probs = [[bar,count*((sample_size))/sample_count/2] for bar, count in freqs.items()]
    keys, vals = zip(*sorted(probs))

    if box:
        ax.bar(keys, vals, width=0.1)
    if f_poly:
        ax.plot(keys, vals, color = "C1")
    if draw_norm:
        X = np.linspace(35,45, 100)
        Y = norm.pdf(X, loc=pop_mean, scale=(pop_var/sample_size)**0.5)
        ax.plot(X, Y, color="C2")

    screen.fill((255,255,255,255))
    screen.blit(plot_to_surface(fig), (0,0))
    pg.display.flip()

    loops += 1
    if not (loops % 10):
        print(f"loop time: {clock.tick()}")
    else:
        clock.tick()

    if loops % 10 == 0:
        sample_size = min(sample_size+1, 12)

    for event in pg.event.get():
        if event.type == pg.QUIT:
            pg.quit()
            running = False
            


